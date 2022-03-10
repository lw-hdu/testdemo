import openpyxl
#利用openpyxl.Workbook()函数创建新的workbook（工作薄）对象，就是创建新的空的Excel文件。
wb=openpyxl.Workbook()
##wb.active就是获取这个工作薄的活动表，通常就是第一个工作表。
sheet=wb.active
#可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"new title"。
sheet.title='new_title'
#把'数据1'赋值给第一个工作表的A1单元格，就是往A1的单元格中写入了'数据1'。
sheet['A1']='数据1'
#将写入的一行内容放在一个列表里，用append()函数插入
#row=['狮子王','烈火英雄']
#sheet.append(row)
#若想将多行数据插入，可以用for循环
rows=[['无主之城','桃花潭水'],['速递与激情','一条狗的使命']]
for i in rows:
    sheet.append(i)
#保存
wb.save('test.xlsx')

###读取excel中数据
#调用openpyxl中的load_workbook()函数，打开文件
wb=openpyxl.load_workbook('test.xlsx')
#获取所有的工作表
sheetname=wb.sheetnames
#print(sheetname)
#获取工作表
sheet=wb['new_title']
A1_value=sheet['A1'].value
print(A1_value)
